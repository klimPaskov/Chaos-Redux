#!/usr/bin/env python3
"""Export the event catalog's authoritative XLSX data sheets to CSV snapshots.

The workbook is the only editable catalog source. This tool overwrites the
three export-only CSV files atomically after a successful read of the Events,
Clusters, and Scenarios sheets.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import sys
import tempfile
from datetime import date, datetime, time
from pathlib import Path
from typing import Dict, List, Optional

try:
	from openpyxl import load_workbook
except ImportError as error:  # pragma: no cover - environment/setup failure
	raise SystemExit("openpyxl is required. Load the workspace spreadsheet dependencies first.") from error


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = REPO_ROOT / "docs" / "spreadsheets" / "chaos_redux_events_catalog.xlsx"
SHEET_EXPORTS = {
	"Events": "chaos_redux_events_catalog.csv",
	"Clusters": "chaos_redux_clusters_catalog.csv",
	"Scenarios": "chaos_redux_scenarios_catalog.csv",
}
# The workbook can retain formatting or stray notes outside the catalog table.
# Keep exports limited to the declared player-facing columns so those cells do
# not turn into phantom CSV fields or blank records.
SHEET_WIDTHS = {
	"Events": 14,
	"Clusters": 7,
	"Scenarios": 6,
}


def _csv_value(value: object) -> str:
	if value is None:
		return ""
	if isinstance(value, (datetime, date, time)):
		return value.isoformat()
	return str(value)


def _sha256(path: Path) -> str:
	digest = hashlib.sha256()
	with path.open("rb") as stream:
		for chunk in iter(lambda: stream.read(1024 * 1024), b""):
			digest.update(chunk)
	return digest.hexdigest()


def _check_formula_cache(workbook_path: Path) -> None:
	formula_workbook = load_workbook(workbook_path, read_only=True, data_only=False)
	try:
		value_workbook = load_workbook(workbook_path, read_only=True, data_only=True)
		try:
			for sheet_name in SHEET_EXPORTS:
				formula_sheet = formula_workbook[sheet_name]
				value_sheet = value_workbook[sheet_name]
				for formula_row, value_row in zip(
					formula_sheet.iter_rows(),
					value_sheet.iter_rows(),
				):
					for formula_cell, value_cell in zip(formula_row, value_row):
						if formula_cell.data_type == "f" and value_cell.value is None:
							raise RuntimeError(
								f"{workbook_path} contains a formula without a cached value at "
								f"{sheet_name}!{formula_cell.coordinate}. Run the workbook "
								"recalculation workflow before exporting."
							)
		finally:
			value_workbook.close()
	finally:
		formula_workbook.close()


def _stage_sheet_export(sheet, target: Path) -> Dict[str, object]:
	temporary_path: Optional[Path] = None
	try:
		width = SHEET_WIDTHS[sheet.title]
		rows = list(sheet.iter_rows(min_col=1, max_col=width, values_only=True))
		while rows and not any(value is not None and str(value) != "" for value in rows[-1]):
			rows.pop()
		with tempfile.NamedTemporaryFile(
			mode="w",
			encoding="utf-8-sig",
			newline="",
			prefix=f".{target.name}.",
			suffix=".tmp",
			dir=target.parent,
			delete=False,
		) as stream:
			temporary_path = Path(stream.name)
			writer = csv.writer(stream, lineterminator="\n")
			row_count = 0
			for row in rows:
				writer.writerow([_csv_value(value) for value in row])
				row_count += 1
			stream.flush()
			os.fsync(stream.fileno())
		return {
			"sheet": sheet.title,
			"target": target,
			"temporary": temporary_path,
			"rows": row_count,
			"columns": width,
		}
	except Exception:
		if temporary_path is not None:
			temporary_path.unlink(missing_ok=True)
		raise


def export_catalog(workbook_path: Path, output_dir: Optional[Path] = None) -> List[Dict[str, object]]:
	"""Export the three catalog data sheets and overwrite their CSV snapshots."""
	workbook_path = workbook_path.resolve()
	output_dir = (output_dir or workbook_path.parent).resolve()
	if not workbook_path.is_file():
		raise FileNotFoundError(f"Workbook not found: {workbook_path}")
	output_dir.mkdir(parents=True, exist_ok=True)
	_check_formula_cache(workbook_path)

	workbook = load_workbook(workbook_path, read_only=True, data_only=True)
	staged: List[Dict[str, object]] = []
	try:
		try:
			for sheet_name, filename in SHEET_EXPORTS.items():
				if sheet_name not in workbook.sheetnames:
					raise KeyError(f"Workbook is missing required sheet: {sheet_name}")
				target = output_dir / filename
				staged.append(_stage_sheet_export(workbook[sheet_name], target))
		except Exception:
			for item in staged:
				temporary = item.get("temporary")
				if isinstance(temporary, Path):
					temporary.unlink(missing_ok=True)
			raise
	finally:
		workbook.close()

	try:
		for item in staged:
			temporary = item["temporary"]
			target = item["target"]
			assert isinstance(temporary, Path)
			assert isinstance(target, Path)
			os.replace(temporary, target)
			item["sha256"] = _sha256(target)
			item.pop("temporary", None)
		return staged
	finally:
		for item in staged:
			temporary = item.get("temporary")
			if isinstance(temporary, Path):
				temporary.unlink(missing_ok=True)


def _parse_args() -> argparse.Namespace:
	parser = argparse.ArgumentParser(
		description="Export Events, Clusters, and Scenarios from the catalog XLSX."
	)
	parser.add_argument(
		"--workbook",
		type=Path,
		default=DEFAULT_WORKBOOK,
		help="Authoritative XLSX workbook. Defaults to the Chaos Redux event catalog.",
	)
	parser.add_argument(
		"--output-dir",
		type=Path,
		default=None,
		help="Directory for the three CSV exports. Defaults to the workbook directory.",
	)
	return parser.parse_args()


def main() -> int:
	args = _parse_args()
	try:
		result = export_catalog(args.workbook, args.output_dir)
	except (OSError, KeyError, RuntimeError) as error:
		print(f"CSV export failed: {error}", file=sys.stderr)
		return 1

	print(
		json.dumps(
			{
				"status": "success",
				"workbook": str(args.workbook.resolve()),
				"exports": [
					{
						"sheet": item["sheet"],
						"path": str(item["target"]),
						"rows": item["rows"],
						"columns": item["columns"],
						"sha256": item["sha256"],
					}
					for item in result
				],
			},
			ensure_ascii=False,
			indent=2,
		)
	)
	return 0


if __name__ == "__main__":
	raise SystemExit(main())
