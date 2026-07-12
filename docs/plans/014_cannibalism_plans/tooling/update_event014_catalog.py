#!/usr/bin/env python3
"""Align Event 014 and SCN-010 with the live implementation/localisation."""

from copy import copy
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[4]
WORKBOOK = ROOT / "docs" / "spreadsheets" / "chaos_redux_events_catalog.xlsx"


EVENT_ROW = [
	14,
	"Cannibalism",
	"Cannibalism begins with evidence recovered from an army at war. Burial parties vanish, ration ledgers are altered, and isolated formations report deliberate cutting and predation within their own ranks. Scarcity may explain the first crimes, but repeated methods suggest that damaged commands are protecting the perpetrators.\n\nThe crisis tests whether military institutions can restore supply, protect witnesses and the dead, and distinguish frightened soldiers from organized killers. Concealment may preserve calm while allowing the pattern to survive. Public terror may hold a front while teaching other units that predation brings rank and protection.",
	"Ritualized Ranks\n\nScattered predation has become a shared ideology of oaths, marks, promotion, and protected membership. No shared headquarters appears in the records.",
	"The Organized Network\n\nCells in separate countries use matching ledgers, routes, prisoner practices, and operational timing. Captured records identify no common headquarters.",
	"Hannibal Lecter Commands\n\nThe concealed command is publicly revealed, and the mature network begins unification under Hannibal Lecter.",
	None,
	None,
	"The World Is the Larder\n\nLecter's host has joined scattered feeding territories and armed kitchens into one command. Roads, farms, prisons, and conquered cities are treated as parts of a single larder, with surviving states left as prey or resistance enclaves.\n\nOrganized consumption becomes a permanent world order. Every surviving government faces the same expanding host, and the network no longer has any reason to hide.\n\nNo Thaw Will Come\n\nLecter's winter host has surrendered its last human restraints to the Wendigo form. Feeding grounds spread with the cold, and conquered communities are folded into a hunger that treats thaw, harvest, and mercy as weaknesses.\n\nAn advancing winter covers the world. The Wendigo command pursues every surviving country until organized human rule is consumed or driven into isolated refuges.",
	"Minor Fire-Once",
	None,
	None,
	"Implemented",
]


SCENARIO_ROW = [
	"SCN-010",
	"The Hunger Lines",
	"Discipline Collapse: A wartime supply crisis has broken discipline inside selected formations. Field Hunger rises while damaged commands attempt containment before predation spreads beyond the first theaters.\n\nRitual Cells: Officer circles and hidden field kitchens have become organized ritual cells. Cult Cohesion is already visible, and several countries may begin with compromised commands.\n\nSilent Islands: Remote ports and island garrisons have fallen quiet behind broken convoy schedules. Communes begin with mature cells, exposed sea routes, and a growing risk of armed island hosts.\n\nWarlord States: Armed host countries emerge from occupied feeding grounds. Each begins with a regional command, an origin doctrine, scavenged stores, and forces raised from the territory it has seized.\n\nConvergence: Several mature host countries answer a common signal. A public convergence warning begins after launch, leaving the world time to destroy the likely hosts and sever their routes before a final authority emerges.",
	"Discipline Collapse, Ritual Cells, Silent Islands, Warlord States, Convergence",
	"Low: A narrow crisis begins with limited territory and forces. Containment remains possible, but delay will strengthen the cells.\n\nMedium: Several formations or theaters enter the crisis. Supply pressure, concealment, and foreign routes will demand an organized response.\n\nHigh: Mature cells and armed hosts begin with strong cohesion, severe command damage, and multiple routes across the map.\n\nMaximum: A broad international network begins with numerous theaters and host countries. Escalation is immediate, but its supply lines, leaders, territories, and convergence routes can still be attacked.",
	"Implemented",
]


def copy_row_style(sheet, source_row: int, target_row: int, columns: int) -> None:
	for column in range(1, columns + 1):
		source = sheet.cell(source_row, column)
		target = sheet.cell(target_row, column)
		if source.has_style:
			target._style = copy(source._style)
		target.number_format = source.number_format
		target.alignment = copy(source.alignment)
		target.protection = copy(source.protection)
	sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height


def main() -> None:
	workbook = load_workbook(WORKBOOK)
	events = workbook["Events"]
	for column, value in enumerate(EVENT_ROW, 1):
		events.cell(15, column).value = value
	events.row_dimensions[15].height = 409.5
	world_end_font = copy(events["I15"].font)
	world_end_font.sz = 9
	events["I15"].font = world_end_font

	scenarios = workbook["Scenarios"]
	copy_row_style(scenarios, 9, 10, 6)
	for column, value in enumerate(SCENARIO_ROW, 1):
		scenarios.cell(10, column).value = value
	scenarios.row_dimensions[10].height = 400
	scenarios.tables["Manual_Scenarios"].ref = "A1:F10"
	for validation in scenarios.data_validations.dataValidation:
		if validation.formula1 == '"Implemented,New,In progress,To Be Reworked,Buggy,Needs Testing"':
			validation.sqref = "F2:F10"
	conditional_ranges = {str(item.sqref) for item in scenarios.conditional_formatting}
	if "F9:F10" not in conditional_ranges:
		for conditional_range in list(scenarios.conditional_formatting):
			if str(conditional_range.sqref) == "F2:F8":
				for rule in scenarios.conditional_formatting._cf_rules[conditional_range]:
					scenarios.conditional_formatting.add("F9:F10", copy(rule))

	workbook.calculation.fullCalcOnLoad = True
	workbook.calculation.forceFullCalc = True
	workbook.calculation.calcMode = "auto"
	workbook.save(WORKBOOK)
	print(f"Updated Event 014 at Events!A15:M15 and SCN-010 at Scenarios!A10:F10 in {WORKBOOK}")


if __name__ == "__main__":
	main()
